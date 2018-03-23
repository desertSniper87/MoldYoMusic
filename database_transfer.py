import sqlite3
from sqlite3 import Error
 
import openpyxl, os

def create_connection(db_file):
    """ create a database connection to a SQLite database """
    try:
        conn = sqlite3.connect("playlist.sqlite3") #??????????
        cur = conn.cursor()
        print(sqlite3.version)
        return conn
    except Error as e:
        print(e)
    # finally:
        # conn.close()

def create_table(conn):
    try:
        # conn.execute("""CREATE TABLE IF NOT EXISTS playlist(
                        # song_id INTEGER PRIMARY KEY,
                        # song_name TEXT,
                        # metal_instrumental INTEGER,
                        # megasync_bool INTEGER,
                        # memorycard_bool INTEGER,
                        # youtube_bool INTEGER default 0)"""
        # )
        conn.execute("""CREATE TABLE IF NOT EXISTS playlist_youtube(
                        song_id INTEGER PRIMARY KEY,
                        youtube_link TEXT)"""
        )
    except Error as e:
        print(e)
 
if __name__ == '__main__':
    conn = create_connection("playlist.sqlite3")

    print("Opening workbook...")
    filename = 'playlist.xlsx'
    wb = openpyxl.load_workbook(filename)

    sheet = wb.get_active_sheet()
    print(sheet)

    print("Reading rows")

    create_table(conn) 

    # for row in range(2, sheet.max_row + 1):
    i = 1
    for row in range(sheet.max_row, 1, -1):
        print(row, sheet['A' + str(row)].value, sheet['A' + str(row)].fill.start_color.index)
        # conn.execute("""INSERT INTO playlist 
                        # (song_id, song_name)
                     # values(?, ?)""", (i, sheet['A' + str(row)].value))
        # conn.execute("""INSERT INTO playlist 
                        # (megasync_bool, memorycard_bool, youtube_bool)
                     # values(?, ?, ?)
                     # WHERE song_id=i""", ())
        conn.execute("""INSERT INTO playlist_youtube
                        (song_id)
                     values(?) """, (i, ))

        i+=1

    conn.commit()
    conn.close()

